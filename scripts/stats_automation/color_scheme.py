import random
from openpyxl import Workbook
from openpyxl.styles import PatternFill
from openpyxl.utils import get_column_letter

# Создаем новый рабочий файл
wb = Workbook()
ws = wb.active

# Названия типов событий
event_types = [
    "Засыпание", "Использование телефона", "Курение",
    "Взгляд в сторону", "Ремень", "Еда",
    "Водитель не в кадре", "Заслон камеры"
]

# Количество групп (по примеру 2)
groups = ["Айрат (2.5)", "Айрат (3.5)"]

# Заголовки
# Первая строка: названия типов событий, каждое занимает 3 ячейки
header_row_1 = [""]  # пустая ячейка для левого верхнего угла
for event in event_types:
    header_row_1.extend([event, "", ""])

ws.append(header_row_1)

# Вторая строка: "Всего", "Подтверждено", "Неопределено" для каждого типа
header_row_2 = [""]  # пустая ячейка
for _ in event_types:
    header_row_2.extend(["Всего", "Подтверждено", "Неопределено"])

ws.append(header_row_2)

# Заполняем данные для каждой группы
# храним исходные значения для сравнения
data_values = []

for group in groups:
    row = [group]
    group_data = []  # для хранения данных этой группы
    for _ in event_types:
        total = random.randint(0, 20)
        confirmed = random.randint(0, total)
        uncertain = total - confirmed
        row.extend([total, confirmed, uncertain])
        group_data.append({
            'total': total,
            'confirmed': confirmed,
            'uncertain': uncertain
        })
    data_values.append(group_data)
    ws.append(row)

# Теперь добавим формулы для подсчета разницы и подсветим ячейки
# Начинаем с третьей строки (индекс 3), так как строки 1 и 2 - заголовки
start_data_row = 3

for row_idx, group_data in enumerate(data_values, start=start_data_row):
    for i, event in enumerate(event_types):
        # позиции колонок для этого события
        col_total = 2 + i * 3 + 1  # +1 потому что в openpyxl индексы с 1
        col_confirmed = col_total + 1
        col_uncertain = col_total + 2

        total_cell = ws.cell(row=row_idx, column=col_total)
        confirmed_cell = ws.cell(row=row_idx, column=col_confirmed)
        uncertain_cell = ws.cell(row=row_idx, column=col_uncertain)

        # Вставляем формулу для разницы (Total - Confirmed) в ячейку "Неопределено"
        formula = f"={get_column_letter(col_total)}{row_idx}-{get_column_letter(col_confirmed)}{row_idx}"
        uncertain_cell.value = formula

        # Проверяем разницу на основе исходных данных
        total_value = group_data[i]['total']
        confirmed_value = group_data[i]['confirmed']
        difference = total_value - confirmed_value

        # Если разница больше 1, подсвечиваем ячейку красным
        if difference > 1:
            red_fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
            uncertain_cell.fill = red_fill

# Сохраняем файл
wb.save(r"C:\Users\username\Downloads\table2event_data.xlsx")